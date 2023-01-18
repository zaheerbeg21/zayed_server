from datetime import date, datetime
import xlwt
from django import template
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.template import loader
from django.urls import reverse
import json
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.models import User

from zayed_university_app.utils import render_to_pdf
from . import models
from .models import *
from zayed_university_app.models import Log

# import pymssql
import pyodbc
from .models import DepartmentAdminUser, Department, UserType
from django.views.decorators.cache import cache_control
from django.contrib.auth import authenticate, login, logout

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.generic import View
import pandas as pd
import openpyxl
from zayed_university_app.models import Log
from django.db.models import Count

from django.shortcuts import render, HttpResponse
from report_app.forms import UploadForm
import xml.etree.ElementTree as ET
from django.http import JsonResponse, QueryDict
import os
import json
import re
import urllib.request
from xml.sax.handler import ContentHandler
from xml.sax import make_parser
import requests
from django.http import QueryDict
from django.utils.datastructures import MultiValueDict

import urllib.request
from pdfminer.high_level import extract_text

from django.utils import timezone


def parsefile(file):
    parser = make_parser()
    parser.setContentHandler(ContentHandler())
    parser.parse(file)


def validateJSON(jsonData):
    try:
        json.loads(jsonData)
    except ValueError as err:
        return False
    return True


# def fetch_url_result(url):


def admin_check(user):
    if user.is_staff or user.is_superuser:
        return True
    return False


def page_not_found(request):
    return render(request, 'home/page-404.html')


def get_custuser(cname, cadmin, cdepart, chk_true):
    global chk_list
    if chk_true:
        chk_list.clear()
        chk_list.append(cname)
        chk_list.append(cadmin)
        chk_list.append(cdepart)
    else:
        # print("in get custuser ", chk_list)
        return chk_list


def get_connection():
    # # for server
    # conn = pymssql.connect(server='192.168.5.79', user='chatboat_sa', password='ch@tb0@t$@',
    # database='zu_chatbot_log_devel')
    conn = pyodbc.connect('Driver=SQL Server;'  ## for local
                          'Server=ZAHEER;'
                          'Database=zu_chatbot_log_dev_july;'
                          'Trusted_Connection=yes;')
    return conn


def daily_wrn_count():
    connection = ''
    """ query data from the log_record table """
    try:

        connection = get_connection()

        cur_wr_ans = connection.cursor()
        cur_rt_ans = connection.cursor()
        cur_no_ans = connection.cursor()

        wr_ans_select_Query = 'SELECT * from dbo.daily_wrong_count;'
        cur_wr_ans.execute(wr_ans_select_Query)
        wr_ans_day = cur_wr_ans.fetchall()

        rt_ans_select_Query = 'SELECT * from dbo.daily_right_count;'
        cur_rt_ans.execute(rt_ans_select_Query)
        rt_ans_day = cur_rt_ans.fetchall()

        no_ans_select_Query = 'SELECT * from dbo.daily_no_count;'
        cur_no_ans.execute(no_ans_select_Query)
        no_ans_day = cur_no_ans.fetchall()

    except pyodbc.Error as error:
        print("Error while fetching data from SQL", error)
    finally:

        if connection:
            cur_wr_ans.close()
            cur_rt_ans.close()
            cur_no_ans.close()

            connection.close()

    return len(wr_ans_day), len(rt_ans_day), len(no_ans_day)


def monthly_wrn_count():
    connection = ''
    """ query data from the log_record table """

    try:

        connection = get_connection()

        cur_wr_ans = connection.cursor()
        cur_rt_ans = connection.cursor()
        cur_no_ans = connection.cursor()

        wr_ans_select_Query = 'SELECT * from dbo.monthly_wrong_count;'
        cur_wr_ans.execute(wr_ans_select_Query)
        wr_ans_month = cur_wr_ans.fetchall()

        rt_ans_select_Query = 'SELECT * from dbo.monthly_right_count;'
        cur_rt_ans.execute(rt_ans_select_Query)
        rt_ans_month = cur_rt_ans.fetchall()

        no_ans_select_Query = 'SELECT * from dbo.monthly_no_count;'
        cur_no_ans.execute(no_ans_select_Query)
        no_ans_month = cur_no_ans.fetchall()

    except pyodbc.Error as error:
        print("Error while fetching data from SQL", error)
    finally:

        if connection:
            cur_wr_ans.close()
            cur_rt_ans.close()
            cur_no_ans.close()

            connection.close()

    return len(wr_ans_month), len(rt_ans_month), len(no_ans_month)


def deptwise_data(dept_name):
    global busy_users, rep_users
    connection = ''
    dept_log_record = ''
    new_users = ''
    try:

        connection = get_connection()
        cursor = connection.cursor()
        cursor1 = connection.cursor()
        cursor2 = connection.cursor()
        cursor3 = connection.cursor()

        SQL_select_Query = "SELECT * FROM fn_deptwise_ans_data('" + dept_name + "') ;"
        cursor.execute(SQL_select_Query)
        log_record = cursor.fetchall()
        dept_log_record = [list(tp) for tp in log_record]

        rep_select_Query = "SELECT TOP(10) * FROM fn_dept_reptd_usr_cnt('" + dept_name + "') ;"
        cursor1.execute(rep_select_Query)
        rep_users = cursor1.fetchall()
        rep_users = [list(tp) for tp in rep_users]

        print("rep_users---->", rep_users)

        busy_select_Query = "SELECT TOP(10) * FROM fn_deptwise_busy_period_cnt('" + dept_name + "');"
        cursor2.execute(busy_select_Query)
        busy_users = cursor2.fetchall()
        busy_users = [list(tp) for tp in busy_users]
        # print("#>-", busy_users)

        new_select_Query = "SELECT * FROM fn_deptwise_new_usr_cnt('" + dept_name + "');"
        cursor3.execute(new_select_Query)
        new_users = cursor3.fetchall()
        new_users = [list(tp) for tp in new_users]
        # print("new_users-deptwise_data> ", new_users)

    except pyodbc.Error as error:
        print("Error while fetching data from SQL", error)

    finally:
        if connection:
            cursor.close()
            connection.close()
            print("SQL connection is closed")

    return dept_log_record, rep_users, busy_users, new_users


def dept_chart(dept_name):
    data, rep_data, busy_data, new_user_data_ = deptwise_data(dept_name)

    event_categories = list()
    wrong_ans_data = list()
    right_ans_data = list()
    no_ans_data = list()

    for dt in data:

        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer description

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():
            for dt in data:
                if i == dt[1].__str__():
                    if dt[0] == 'Right Answer': temp[0] = dt[2]
                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]
                    if dt[0] == 'No Answer': temp[2] = dt[2]
        edict[i] = temp

    # print(edict)

    for key, value in edict.items():
        right_ans_data.append(value[0])
        wrong_ans_data.append(value[1])
        no_ans_data.append(value[2])

    wrong_answer = {
        'name': 'wrong_answer',
        'data': wrong_ans_data,
        'color': '#e53935'
    }

    right_answer = {
        'name': 'right_answer',
        'data': right_ans_data,
        'color': '#43a047'
    }

    no_answer = {
        'name': 'no_answer',
        'data': no_ans_data,
        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    return bar_dump, line_dump, rep_data, busy_data, new_user_data_


def get_repeated_bot_users():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT TOP(10) * from dbo.repeated_bot_users_day_view order by ev_count desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_repeated_bot_monthly_users():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT TOP(10) * from dbo.repeated_bot_users_monthly_view order by ev_count desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)

    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_busy_period_count():
    busy_count = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT TOP(10) * from dbo.busy_period_count order by intent_counts desc;'
        cursor.execute(SQL_select_Query)

        busy_count = cursor.fetchall()

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count


def get_busy_period_count_monthly():
    busy_count_month = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        SQL_select_Query = 'SELECT TOP(10) * from dbo.busy_period_count_monthly;'
        cursor.execute(SQL_select_Query)

        busy_count_month = cursor.fetchall()
        # print("TOP 10 busy_count_month", busy_count_month)

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)

    finally:

        if connection:
            cursor.close()

            connection.close()

            # print("SQL connection is closed")

    return busy_count_month


def live_count():
    live = []
    live_month = []
    connection = ''
    try:
        connection = get_connection()
        cursor = connection.cursor()

        cursor1 = connection.cursor()

        SQL_select_Query = 'Select Top 1 * from dbo.livechat_daywise order by day desc;'
        cursor.execute(SQL_select_Query)

        live = cursor.fetchall()
        SQL_select_Query1 = 'Select top 1 * from dbo.livechat_monthwise order by month desc;'
        cursor1.execute(SQL_select_Query1)

        live_month = cursor1.fetchall()

    except pyodbc.Error as error:
        print(error)

    finally:
        if connection:
            cursor.close()
            connection.close()

    return live, live_month


def reset_count():
    reset_count = []

    reset_month = []
    connection = ''
    try:

        connection = get_connection()

        cursor = connection.cursor()

        cursor1 = connection.cursor()

        # SQL_select_Query = 'Select Top 1 * from dbo.reset_daywise order by day desc;' # emergency exit
        SQL_select_Query = 'Select  * from dbo.daily_reset_count;'
        cursor.execute(SQL_select_Query)

        reset_count = cursor.fetchall()

        # print("reset_count", reset_count)

        SQL_select_Query1 = 'Select * from dbo.monthly_reset_count;'
        cursor1.execute(SQL_select_Query1)

        reset_month = cursor1.fetchall()

        # print("reset_month", reset_month)

    except pyodbc.Error as error:

        print("Error while fetching data from SQL server", error)

    finally:

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    return len(reset_count), len(reset_month)


def get_total_users_cnt():
    total_eng_users_cnt = 0

    total_new_users_cnt = 0
    connection = ''
    """ query data from the log_record table """

    try:

        connection = get_connection()

        cursor2 = connection.cursor()

        cursor1 = connection.cursor()

        # postgreSQL_select_Query = 'SELECT * from chatbot.total_cnt_view_mview;'

        eng_select_Query = 'SELECT * from dbo.engaged_users_view1;'

        cursor2.execute(eng_select_Query)

        new_select_Query = 'SELECT * from dbo.new_users_view1;'

        cursor1.execute(new_select_Query)

        # log_record = cursor.fetchone()

        total_eng_users_cnt = cursor2.fetchall()

        total_new_users_cnt = cursor1.fetchall()

        # print("total_eng_users_cnt---> ", total_eng_users_cnt)
        # #
        # print("total_new_users_cnt---> ", total_new_users_cnt)

        # total_new_users_cnt = [list(tp) for tp in desc_total_cnt]



    except pyodbc.Error as error:

        print("Error while fetching total_desc_cnt from SQL", error)



    finally:

        # closing database connection.

        if connection:
            cursor2.close()

            cursor1.close()

            connection.close()

            print("SQL connection is closed")

    return total_eng_users_cnt, total_new_users_cnt


def get_monthly_data():
    list_log_record = ''
    connection = ''
    """ query data from the log_record table """

    try:

        connection = get_connection()

        cursor = connection.cursor()

        cur_month_eng = connection.cursor()

        cur_month_new = connection.cursor()

        # select_Query = 'SELECT * from dbo.monthly_data_view;'
        # cursor.execute(select_Query)
        # log_record = cursor.fetchall()
        # list_log_record = [list(tp) for tp in log_record]

        wr_answer_cnt = Log.objects.all().filter(user_datetime__month=datetime.now().month, event_type_id='3').count()
        rt_answer_cnt = Log.objects.all().filter(user_datetime__month=datetime.now().month, event_type_id='4').count()
        no_answer_cnt = Log.objects.all().filter(user_datetime__month=datetime.now().month, event_type_id='5').count()
        
        log_record = [('Right Answer', datetime.now().month, rt_answer_cnt),
                      ('No Answer', datetime.now().month, no_answer_cnt),
                      ('Wrong Answer', datetime.now().month, wr_answer_cnt)]

        # # print(">>>>>>>>>>>>>", log_record)
        list_log_record = [list(tp) for tp in log_record]

        # eng_select_Query = 'SELECT * from dbo.monthly_engaged_users_view;'
        # cur_month_eng.execute(eng_select_Query)
        # count_eng_usr_month = cur_month_eng.fetchall()

        new_select_Query = 'SELECT * from dbo.monthly_new_users_view;'
        cur_month_new.execute(new_select_Query)
        new_usr_month = cur_month_new.fetchall()

        count_eng_usr_month = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(
            user_datetime__month=datetime.now().month,event_type_id=1,
            login_count__gt=1)


        # _usr = Log.objects.values('user_email').annotate(login_count=Count('event_type_id')).filter(
        #     user_datetime__month=datetime.now().month)
        # new_usr_month = Log.objects.filter(user_email__in=[item['user_email'] for item in _usr],
        #                                    user_datetime__month=datetime.now().month)

        new_usr_month = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(
            user_datetime__month=datetime.now().month,event_type_id=1,
            login_count=1)
   

        print("-------------------", count_eng_usr_month)
        print("*******************>", new_usr_month)
        # print("into get_monthly_data = > ", list_log_record)


    except pyodbc.Error as error:

        print("Error while fetching data from SQL", error)



    finally:

        if connection:
            cursor.close()

            connection.close()

    return list_log_record, count_eng_usr_month, new_usr_month


def get_daily_data():
    engaged_u = 0

    new_u = 0
    connection = ''
    list_log_record = ''

    """ query data from the log_record table """

    try:
        connection = get_connection()

        cursor = connection.cursor()

        # SQL_select_Query = 'SELECT * from dbo.daily_data_view;'
        # cursor.execute(SQL_select_Query)
        # log_record = cursor.fetchall()
        # # print("in daily data- ", log_record)
        # list_log_record = [list(tp) for tp in log_record]

        cur_eng_usr = connection.cursor()

        cur_new_usr = connection.cursor()

        wr_answer_cnt = Log.objects.all().filter(user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59), event_type_id='3').count()
        rt_answer_cnt = Log.objects.all().filter(user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59), event_type_id='4').count()
        no_answer_cnt = Log.objects.all().filter(user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59), event_type_id='5').count()
        
        log_record = [('Right Answer', datetime.now().date(), rt_answer_cnt),
                      ('No Answer', datetime.now().date(), no_answer_cnt),
                      ('Wrong Answer', datetime.now().date(), wr_answer_cnt)]
        list_log_record = [list(tp) for tp in log_record]

        # eng_select_Query = 'SELECT * from dbo.daily_eng_users_view;'
        # cur_eng_usr.execute(eng_select_Query)

        # count_eng_users = cur_eng_usr.fetchall()

        # new_select_Query = 'SELECT * from dbo.daily_new_users_view;'
        # cur_new_usr.execute(new_select_Query)
        # new_users = cur_new_usr.fetchall()

        # count_eng_users = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(
        #     user_datetime__date=datetime.now().date(),
        #     login_count=2)
        
        # _usr = Log.objects.values('user_email').annotate(login_count=Count('event_type_id')).filter(login_count=1)
        # new_users = Log.objects.filter(user_email__in=[item['user_email'] for item in _usr],
        #                                user_datetime__date=datetime.now().date())
        # print("-------------------", count_eng_users)
        # print("*******************>", new_users)

        count_eng_users = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(
        event_type_id=1,login_count__gt=1, user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59))


        new_users = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(
            event_type_id=1,user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59),
            login_count=1)

        print("-------------------", count_eng_users)
        print("*******************>", new_users)

        # self.filter(location=location_id, start_gte=timezone.now().replace(hour=0, minute=0, second=0), end_lte=timezone.now().replace(hour=23, minute=59, second=59))

    except pyodbc.Error as error:

        print("Error while fetching data from SQL", error)

    finally:

        # closing database connection.

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    # engaged_u, new_u =  get_total_users_cnt()
    return list_log_record, count_eng_users, new_users


def get_total_event_cnt():
    global total_desc_cnt

    connection = ''

    try:

        connection = get_connection()

        cursor = connection.cursor()

        # postgreSQL_select_Query = 'SELECT * from chatbot.total_cnt_view_mview;'

        select_Query = 'SELECT * from dbo.total_ans_cnt_view;'

        cursor.execute(select_Query)

        # print("Selecting rows from chatbot_log table using cursor.fetchall")

        # log_record = cursor.fetchone()

        desc_total_cnt = cursor.fetchall()

        total_desc_cnt = [list(tp) for tp in desc_total_cnt]

    except pyodbc.Error as error:

        print("Error while fetching total_desc_cnt from SQL", error)

    finally:

        # closing database connection.

        if connection:
            cursor.close()

            connection.close()

            print("SQL connection is closed")

    return total_desc_cnt


@csrf_protect
def daily_charts(request):
    users_list = User.objects.filter(is_superuser=False)

    data, eng_usr_daily, new_usr_daily = get_daily_data()
    print("data, eng_usr_daily, new_usr_daily ", data, eng_usr_daily, new_usr_daily)
    #     wr_day, rt_day, no_day = daily_wrn_count()
    tot_ans_cnt = get_total_event_cnt()
    # print("tot_ans_cnt_wr =============== ", tot_ans_cnt[0])

    busy_count = get_busy_period_count()

    busy_count_list = []

    for i in busy_count:

        if len(i[0]) > 5:
            busy_count_list.append(i)

    repeated_users = get_repeated_bot_users()

    #     reset, reset_month = reset_count()
    #     live, live_month = live_count()
    _, reset_month = reset_count()
    _, live_month = live_count()

    reset = Log.objects.all().filter(user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59), event_type_id='7').count()
    live = Log.objects.all().filter(user_datetime__gte=timezone.now().replace(hour=0, minute=0, second=0),user_datetime__lte=timezone.now().replace(hour=23, minute=59, second=59), event_type_id='6').count()

    # print("""================================= in daily charts =================================""")
    # total_usr_daily = Log.objects.filter(user_datetime__date=datetime.now().date()).values(
    #     "user_email").distinct().count()
    # # Visit.objects.filter(stuff).values("ip_address").distinct().count()
    # print(">>> ", total_usr_daily)

    event_categories = list()

    wrong_ans_data = list()

    right_ans_data = list()

    no_ans_data = list()

    # print("data---", data)
    for dt in data:
        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer desc

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():

            for dt in data:

                if i == dt[1].__str__():

                    if dt[0] == 'Right Answer': temp[0] = dt[2]

                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]

                    if dt[0] == 'No Answer': temp[2] = dt[2]

        edict[i] = temp

    for key, value in edict.items():
        right_ans_data.append(value[0])

        wrong_ans_data.append(value[1])

        no_ans_data.append(value[2])

    wrong_answer = {

        'name': 'wrong_answer',

        'data': wrong_ans_data,

        'color': '#e53935'

    }

    right_answer = {

        'name': 'right_answer',

        'data': right_ans_data,

        'color': '#43a047'

    }

    no_answer = {

        'name': 'no_answer',

        'data': no_ans_data,

        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Chatbot Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Chatbot Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Date',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    # for d in reset:

    #     print("---", d[1])

    # busy_count.remove(i)

    # print("busy_count_list", busy_count_list)
    # print('reset', reset)

    context = {

        # 'reset_count': len(reset) if len(reset) == 0 else reset[0][1],
        # 'live_count': len(live) if len(live) == 0 else live[0][1],

        'reset_count': reset,
        'live_count': live,

        'repeated_users': repeated_users,

        'busy_count': busy_count_list,

        'bar_chart': bar_dump,

        'line_chart': line_dump,

        'no_answer': "No Answer", 'no_answer_count': no_ans_data[0],

        'rt_answer': "Right Answer", 'rt_answer_count': right_ans_data[0],

        'wr_answer': "Wrong Answer", 'wr_answer_count': wrong_ans_data[0],

        'total_users': len(eng_usr_daily) + len(new_usr_daily),

        'engaged_users': len(eng_usr_daily),

        'new_users': len(new_usr_daily),

        'users_list': users_list

    }

    return context


def getUserById(_userid):
    users = User.objects.get(id=_userid)

    # print(users)


@csrf_protect
def monthly_charts(request):
    repeated_users = []

    reset, reset_month = [], []

    data, eng_usr_monthly, new_usr_monthly = get_monthly_data()
    #     wr_mon, rt_mon, no_mon = monthly_wrn_count()
    tot_ans_cnt = get_total_event_cnt()

    _, live_month = live_count()

    repeated_users = get_repeated_bot_monthly_users()

    busy_period_count_month = get_busy_period_count_monthly()

    event_categories = list()

    wrong_ans_data = list()

    right_ans_data = list()

    no_ans_data = list()

    reset, reset_month = reset_count()

    # print("""================================= in monthly charts =================================""")
    for dt in data:

        if dt[1].__str__() not in event_categories:
            event_categories.append(dt[1].__str__())  # for answer desc

    edict = {}

    for i in event_categories:

        temp = [0, 0, 0]

        if i not in edict.keys():

            for dt in data:

                if i == dt[1].__str__():

                    if dt[0] == 'Right Answer': temp[0] = dt[2]

                    if dt[0] == 'Wrong Answer': temp[1] = dt[2]

                    if dt[0] == 'No Answer': temp[2] = dt[2]

        edict[i] = temp

    for key, value in edict.items():
        right_ans_data.append(value[0])

        wrong_ans_data.append(value[1])

        no_ans_data.append(value[2])

    # print("wr month", wrong_ans_data[1])
    # print("rt month", right_ans_data[0])
    # print("no month", no_ans_data[0])

    wrong_answer = {

        'name': 'wrong_answer',

        'data': wrong_ans_data,

        'color': '#e53935'

    }

    right_answer = {

        'name': 'right_answer',

        'data': right_ans_data,

        'color': '#43a047'

    }

    no_answer = {

        'name': 'no_answer',

        'data': no_ans_data,

        'color': '#fb8c00'

    }

    bar_chart = {

        'chart': {'type': 'column',

                  },

        'title': {'text': 'Chatbot Log Summary(Bar Chart)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Month',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    line_chart = {

        'chart': {'type': 'line'},

        'title': {'text': 'Chatbot Log Summary(Line Charts)',

                  'style': {

                      ' color': '#000',

                      'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif',

                  }

                  },

        'xAxis': {'categories': event_categories,

                  'title': {

                      'text': 'Month',

                      ' align': 'high'

                  }

                  },

        'yAxis': {

            'title': {

                'text': 'Count',

                ' align': 'high',

                'style': {

                    ' color': '#000',

                    'font': 'bold 16px "Trebuchet MS", Verdana, sans-serif'

                }

            }

        },

        'series': [wrong_answer, right_answer, no_answer],

        'legend': {

            'itemStyle': {

                'font': '9pt Trebuchet MS, Verdana, sans-serif',

                'color': 'black'

            },

            'itemHoverStyle': {

                'color': 'red'

            },

        }

    }

    bar_dump = json.dumps(bar_chart)

    line_dump = json.dumps(line_chart)

    temp_busy_count = []

    for i in busy_period_count_month:
        if i[0] != "": temp_busy_count.append(i)

    # print('[live_month info]', live_month)
    # print("bar_chart ", bar_dump)
    context = {

        # 'reset_month': len(reset_month) if len(reset_month) == 0 else reset_month[0][1],
        'reset_month': reset_month,
        'repeated_users': repeated_users,

        'live_count': len(live_month) if len(live_month) == 0 else live_month[0][1],

        'busy_period_count_month': temp_busy_count,

        'total_users': len(eng_usr_monthly) + len(new_usr_monthly),

        'new_users': len(new_usr_monthly),

        'engaged_users': len(eng_usr_monthly),

        'bar_chart': bar_dump,

        'line_chart': line_dump,

                # 'no_answer': "No Answer", 'no_answer_count': no_mon,

                # 'rt_answer': "Right Answer", 'rt_answer_count': rt_mon,

                # 'wr_answer': "Wrong Answer", 'wr_answer_count':   wr_mon,
        'no_answer': "No Answer", 'no_answer_count': no_ans_data[0],

        'rt_answer': "Right Answer", 'rt_answer_count': right_ans_data[0],

        'wr_answer': "Wrong Answer", 'wr_answer_count': wrong_ans_data[0],

    }
    return context


def create_excel(exl_name, col_desc, col_lst, request, rows):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Report.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(exl_name, cell_overwrite_ok=True)
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    font_style.font.bold = True
    cols_tab_desc = col_desc
    cols_head = col_lst
    font_style = xlwt.XFStyle()
    for col_num in range(len(cols_tab_desc)):
        ws.write(row_num, col_num, cols_tab_desc[col_num], font_style)

    for col_num in range(len(cols_head)):
        ws.write(row_num, col_num, cols_head[col_num], font_style)

    # print("================= # =========================", type(context['repeated_users'][0]))
    context = monthly_charts(request)
    rows = rows
    for row in rows:
        row_num += 1
        for row[0] in row:
            # print("================= # =========================",row[0])
            row_num += 1
            for col_num in range(len(row[0])):
                ws.write(row_num, col_num, str(row[0][col_num]), font_style)
                # print("WOrked !!!")
    wb.save(response)

    if wb:
        return response
    return HttpResponse("No Data Found.")


@login_required
@user_passes_test(admin_check)
def report(request, report_name):
    # print('==============================', request.path)

    if request.path == '/export_pdf/Daily/':
        context = daily_charts(request)
        pdf = render_to_pdf('home/daily_export.html', context)
        if pdf:
            return HttpResponse(pdf, content_type='application/pdf')
        return HttpResponse("PDF Not Found.")
    elif request.path == '/export_pdf/Monthly/':
        context = monthly_charts(request)
        pdf = render_to_pdf('home/monthly_export.html', context)
        if pdf:
            return HttpResponse(pdf, content_type='application/pdf')
        return HttpResponse("PDF Not Found.")

    elif request.path == '/export_excel/Daily/':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="DailyReport.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Report', cell_overwrite_ok=True)
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Total Users', 'Engaged Users', 'New Users', 'Reset Count', 'Live Chat', 'No Answer',
                   'Right Answer', 'Wrong Answer']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        font_style = xlwt.XFStyle()

        context = daily_charts(request)
        # print("context => ", context)

        rows_stats = [(context['total_users'], context['engaged_users'], context['new_users'], context['reset_count'],
                       context['live_count'], context['no_answer_count'], context['rt_answer_count'],
                       context['wr_answer_count'])]

        for row in rows_stats:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)

        font_style.font.bold = True
        cols_tab = ['Table for Repeated Interacted Users Data']
        cols_repeated_users = ['User Email', 'Bot Interaction', 'Date Time']

        for col_num in range(len(cols_tab)):
            ws.write(3, col_num, cols_tab[col_num], font_style)

        for col_num in range(len(cols_repeated_users)):
            ws.write(4, col_num, cols_repeated_users[col_num], font_style)

        # print("================= # =========================", type(context['repeated_users'][0]))

        rows_repeated_users = [(context['repeated_users'])]
        for row in rows_repeated_users:
            row_num += 3
            for row[0] in row:
                # print("================= # =========================",row[0])
                row_num += 1
                for col_num in range(len(row[0])):
                    ws.write(row_num, col_num, str(row[0][col_num]), font_style)

        font_style.font.bold = True
        # busiest period table
        cols_tab = ['Table for Busiest period of Chatbot interaction']
        cols_busy_period_count = ['Department', 'Count', 'Date Time']

        for col_num in range(len(cols_tab)):
            ws.write(17, col_num, cols_tab[col_num], font_style)

        for col_num in range(len(cols_busy_period_count)):
            ws.write(18, col_num, cols_busy_period_count[col_num], font_style)

        rows_busy_prd_cnt_daily = [(context['busy_count'])]
        # print(context['busy_count'])
        # print("================= # =========================", type(context['busy_count'][0]))

        for row in rows_busy_prd_cnt_daily:
            row_num += 4
            for row[0] in row:
                # print("================= # =========================",row[0])
                row_num += 1
                for col_num in range(len(row[0])):
                    ws.write(row_num, col_num, str(row[0][col_num]), font_style)

        wb.save(response)
        if wb:
            return response
        return HttpResponse("No Data Found.")

    elif request.path == '/export_excel/Monthly/':
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="MonthlyReport.xls"'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Report', cell_overwrite_ok=True)
        row_num = 0
        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = ['Total Users', 'Engaged Users', 'New Users', 'Reset Count', 'Live Chat', 'No Answer',
                   'Right Answer', 'Wrong Answer']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        font_style = xlwt.XFStyle()

        context = monthly_charts(request)
        # print("context => ", context)

        rows_stats = [(context['total_users'], context['engaged_users'], context['new_users'], context['reset_month'],
                       context['live_count'], context['no_answer_count'], context['rt_answer_count'],
                       context['wr_answer_count'])]

        for row in rows_stats:
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)

        font_style.font.bold = True
        cols_tab = ['Table for Repeated Interacted Users Data']
        cols_repeated_users = ['User Email', 'Bot Interaction', 'Date Time']

        for col_num in range(len(cols_tab)):
            ws.write(3, col_num, cols_tab[col_num], font_style)

        for col_num in range(len(cols_repeated_users)):
            ws.write(4, col_num, cols_repeated_users[col_num], font_style)

        rows_repeated_users = [(context['repeated_users'])]
        for row in rows_repeated_users:
            row_num += 3
            for row[0] in row:
                # print("================= # =========================",row[0])
                row_num += 1
                for col_num in range(len(row[0])):
                    ws.write(row_num, col_num, str(row[0][col_num]), font_style)

        font_style.font.bold = True
        # busiest period table
        cols_tab = ['Table for Busiest period of Chatbot interaction']
        cols_busy_period_count = ['Department', 'Count', 'Date Time']

        for col_num in range(len(cols_tab)):
            ws.write(17, col_num, cols_tab[col_num], font_style)

        for col_num in range(len(cols_busy_period_count)):
            ws.write(18, col_num, cols_busy_period_count[col_num], font_style)

        rows_busy_prd_cnt_month = [(context['busy_period_count_month'])]
        # print(context['busy_period_count_month'])
        # print("================= # =========================", type(context['busy_period_count_month'][0]))

        for row in rows_busy_prd_cnt_month:
            row_num += 4
            for row[0] in row:
                # print("================= # =========================",row[0])
                row_num += 1
                for col_num in range(len(row[0])):
                    ws.write(row_num, col_num, str(row[0][col_num]), font_style)

        wb.save(response)

        if wb:
            return response
        return HttpResponse("No Data Found.")

    else:

        if report_name == "Daily":

            context = daily_charts(request)

            return render(request, 'home/user_report.html', context)

        elif report_name == "Monthly":
            # print("into elif before monthly charts", report_name)

            context = monthly_charts(request)
            # print("in reports context - > ", context)

            return render(request, 'home/month_report.html', context)


@login_required
@user_passes_test(admin_check)
def users(request):
    reports = Report.objects.values("assigned_to", "report_name")
    users_list = User.objects.filter(is_superuser=False)

    dept_usr_list = DepartmentAdminUser.objects.values("user", "usertype", "department")
    usertype_lst = DepartmentAdminUser.objects.all()  # added
    department = Department.objects.all()  # added

    ulist = []
    for u in usertype_lst:
        if u.user.username not in ulist:
            ulist.append(u.user.username)
            ulist.append(u.usertype.usertype)

    temp_list = []
    temp_dp_list = []
    for dpt in department:
        # print("dpt=> ", dpt.department)
        if dpt.department not in temp_dp_list:
            temp_dp_list.append(dpt.id)
            temp_dp_list.append(dpt.department)
            temp_list.append(dpt.department)

    # for d in dept_usr_list:
    #     print("-?>", d.department)

    _user_reports = {}
    for user in users_list:
        _temp = []
        for dptl in dept_usr_list:
            # print("Dictionary-", dptl)
            if user.id == dptl['user']:
                # print("################################")
                # print("", user.id, dptl['user'])
                # print("", dptl['department'])
                # print(type(dptl['department']))
                # print("", temp_dp_list.index(dptl['department']))
                # print("", temp_dp_list.index(dptl['department']) + 1)

                _temp.append(temp_dp_list[temp_dp_list.index(dptl['department']) + 1])

        if user.id not in _user_reports:
            _user_reports[user.id] = _temp
    # print("user_reports - ", _user_reports)
    if len(users_list) == 0:
        html = ""
        html += '<tr><td colspan=3>No Users Found</td></tr>'
        context = {
            'report_html': html,
            'depart_name': request.session['depart']
        }
        return render(request, 'home/user_admin.html', context)
    else:
        html = ""
        for user in range(len(users_list)):
            ulist_var = ulist[ulist.index(users_list[user].username) + 1]
            html += '<tr><td>' + str(user + 1) + '</td><td>' + users_list[
                user].username + '<td>' + str(ulist_var) + '</td>' + '</td><td><button class="dropbtn" id="' + str(
                user + 1) + '" onclick="show(' + str(
                user + 1) + ')" value="Reports Assigned">Departments Assigned</button></td>'
            for tmp in temp_list:
                html += '<tr style="text-align:center" class="show' + str(
                    user + 1) + '" id="reports_hide"><td  colspan="3">' + tmp + '</td>'
                inserted = False
                for key, value in _user_reports.items():
                    # print("key value ",  key, value)
                    if users_list[user].id == key:
                        for i in value:
                            if i == tmp:
                                inserted = True
                                html += '<td><input type="checkbox" style="margin-right: 180px;" value="assigned" onclick=check(event,' + str(
                                    users_list[user].id) + ',"' + tmp + '") checked/></td>'
                if not inserted:
                    html += '<td><input type="checkbox" style="margin-right: 180px;" value="assigned" onclick=check(event,' + str(
                        users_list[user].id) + ',"' + tmp + '") /></td>'
                html += '</tr>'
            html += '</tr>'
        context = {
            'reports': temp_list,
            'user_list': users_list,
            'report_html': html,
            'depart_name': request.session['depart']
        }
        return render(request, 'home/user_admin.html', context)


def update_report(request):
    # print('############## update_report #################')
    # print("1.", request.POST)
    # print("2.", request.POST['userid'])
    # print("3.", request.POST['reportname'])
    # print("4.", request.POST['reportname'][0])
    userid_ = User.objects.get(id=request.POST['userid'])
    userid_name = userid_.username
    # print("userid_ ", userid_)
    # print("userid_username  ", userid_name)

    ulist = []
    usertype_lst = DepartmentAdminUser.objects.all()
    for u in usertype_lst:
        if u.user.username not in ulist:
            ulist.append(u.user.username)
            ulist.append(u.usertype.usertype)

    u_type = ulist[ulist.index(userid_name) + 1]
    # print("ulist - ", ulist)
    # print("utype - ", u_type)
    # print("utype type - ", type(u_type))

    report_nm = request.POST["reportname"]
    # print("report_nm - ", report_nm)
    # print("type - ", type(report_nm))
    did = Department.objects.get(department=report_nm)
    uid = UserType.objects.get(usertype=u_type)
    # print("did ================> ", did)
    # print("uid ================> ", uid)
    # userid_, uid, did,

    if request.POST['checked'] == 'true':
        DepartmentAdminUser.objects.create(user=userid_, usertype=uid, department=did)
    else:
        instance = DepartmentAdminUser.objects.get(user=userid_, usertype=uid, department=did)
        # print(">>> ", instance)
        instance.delete()
    return redirect('user/')


@csrf_protect
@login_required
@user_passes_test(admin_check)
def index(request):
    current_user = request.user
    # print(current_user.id)
    # print(current_user.username)
    users_list = User.objects.filter(is_superuser=False)

    if current_user.username == "admin":

        reports = Report.objects.values("assigned_to", "report_name")
        departments = Department.objects.all()

        temp_list = []

        for i in reports:

            if i["report_name"] not in temp_list:
                temp_list.append(i["report_name"])

        return render(request, 'home/index_admin.html', {'userlist': users_list, 'reports': temp_list,
                                                         'depts': departments})
        # 'depart_name': request.session['depart']})

    else:

        reports = Report.objects.values("assigned_to", "report_name")

        temp_list = []

        for i in reports:

            if i["report_name"] not in temp_list and i['assigned_to'] == current_user.id:
                temp_list.append(i["report_name"])

        return render(request, 'home/user_adoreta.html', {'reports': temp_list})
        # {'reports': temp_list, 'depart_name': request.session['depart']})


@csrf_protect
@login_required(login_url="/login/")
def pages(request):
    context = {}

    try:

        load_template = request.path.split('/')[-1]

        # print('load_template', load_template)

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))

        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)

        return HttpResponse(html_template.render(context, request))



    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')

        return HttpResponse(html_template.render(context, request))

    except:

        html_template = loader.get_template('home/page-500.html')

        return HttpResponse(html_template.render(context, request))


@csrf_protect
@login_required(login_url="/admin/")
def index_admin(request):
    context = {'segment': 'index_admin'}

    html_template = loader.get_template('home/index_admin.html')

    return HttpResponse(html_template.render(context, request))


@csrf_protect
@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_register(request):
    session_dept = request.session['depart']
    print("session_dept - ", session_dept)
    reg_user = request.user
    if reg_user is not None:
        name_title = reg_user.username
        chk_super = User.objects.filter(username=name_title, is_superuser=True).exists()
        if chk_super:
            depart_name = "Super_Admin"
            admin_type = 'admin'
        else:
            did = Department.objects.get(department=session_dept)
            user_ = DepartmentAdminUser.objects.get(user=reg_user, department=did)
            admin_type = user_.usertype.usertype
            depart_name = user_.department.department

            # print("admin_type => ", admin_type)
            # print("depart_name ", depart_name)
            # print("type -admin ", type(admin_type))

            if admin_type == 'Department_User':
                # print("=inside if ", admin_type)
                return redirect('page_not_found')

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm = request.POST.get('confirm')
        utype = request.POST.get('utype')
        dtype = request.POST.getlist('dtype')
        # print("dtype-> ", dtype)
        if len(dtype) == 0:
            dtype = [session_dept]

        list_usertype = UserType.objects.all()
        list_dept = Department.objects.all()
        # list_email = Department.objects.all().values('user_email')
        if password != confirm:
            color = "red"
            context = {
                'color': color,
                'depart_name': depart_name,
                'name_title': name_title,
                'admin_type': admin_type,
                'msg': 'PASSWORD & CONFIRM PASSWORD DID NOT MATCH !',
                'list_usertype': list_usertype,
                'list_dept': list_dept,
            }
            return render(request, 'accounts/register_.html', context)
        else:
            # print("department=> ", dtype)

            # create django username
            user = User.objects.create(username=username, email=email)
            user.set_password(password)
            user.save()

            for dt in dtype:
                did = Department.objects.get(department=dt)
                uid = UserType.objects.get(usertype=utype)
                dau = DepartmentAdminUser.objects.create(user=user, usertype=uid, department=did)

                # print("dt========", dt, did, username, email, password, utype)
                dau.save()

            # print("User Created!!!")
            color = 'green'
            context = {
                'color': color,
                'depart_name': depart_name,
                'name_title': name_title,
                'admin_type': admin_type,
                'msg': f'ACCOUNT CREATED FOR {username} !',
                'list_usertype': list_usertype,
                'list_dept': list_dept,
            }
            return render(request, 'accounts/register_.html', context)

    list_usertype = UserType.objects.all()
    list_dept = Department.objects.all()

    context = {
        'depart_name': depart_name,
        'name_title': name_title,
        'admin_type': admin_type,
        'list_usertype': list_usertype,
        'list_dept': list_dept,
    }

    return render(request, 'accounts/register_.html', context)


@login_required
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def get_dept_data(request, department_name):
    dept_list = Department.objects.all()
    updated_list = []

    for dl in dept_list:
        updated_list.append(dl.department)

    if department_name not in updated_list:
        return redirect('page_not_found')
    else:
        depart_user = request.user
        if depart_user.is_superuser:
            name_title = depart_user.username
            depart_name = "Super_Admin"
            admin_type = "admin"
        else:
            did = Department.objects.get(department=department_name)
            user_ = DepartmentAdminUser.objects.filter(user=depart_user, department=did)
            # print("user_ = > ", user_)
            if len(user_) == 0:
                return redirect('page_not_found')

            name_title = depart_user.username
            depart_name = user_[0].department.department
            admin_type = user_[0].usertype.usertype

        wr_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='3').count()
        rt_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='4').count()
        no_answer_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='5').count()
        live_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='6').count()
        reset_cnt = Log.objects.filter(intent=department_name).filter(event_type_id='7').count()
        total_users_cnt = Log.objects.all().values('user_email').distinct().filter(intent=department_name).count()
        # eng_users_cnt = Log.objects.all().values('user_email').filter(intent=department_name).count()

        repeated_users_list = Log.objects.all().values('user_email', 'user_datetime', 'intent')

        # _usr = Log.objects.values('user_email').annotate(login_count=Count('user_email')).filter(intent=department_name,
        #                                                                                          login_count=1)
        #
        # dept_new_users = Log.objects.filter(user_email__in=[item['user_email'] for item in _usr],
        #                                     intent=department_name, event_type_id='1')

        _usr = Log.objects.values('user_email').annotate(login_count=Count('event_type_id')).filter(
            intent=department_name)
        dept_new_users = Log.objects.filter(user_email__in=[item['user_email'] for item in _usr],
                                            intent=department_name, event_type_id='1')

        ## bar and line chart

        bar_data, line_data, rep_data, busy_data, new_user_data__ = dept_chart(department_name)
        # print("===", total_users_cnt , len(new_user_data__))
        eng_users_cnt = total_users_cnt - len(dept_new_users)
        context = {
            'admin_type': admin_type,
            'depart_name': depart_name,
            'name_title': name_title,
            'dept_name': department_name,
            'wr_answer': 'Wrong Answer', 'wr_answer_count': wr_answer_cnt,
            'rt_answer': 'Right Answer', 'rt_answer_count': rt_answer_cnt,
            'no_answer': 'No Answer', 'no_answer_count': no_answer_cnt,
            'live_count': live_cnt,
            'reset_count': reset_cnt,
            'total_users': total_users_cnt,
            'eng_users_cnt': eng_users_cnt,
            'repeated_users_list': repeated_users_list,
            'bar_chart_data': bar_data,
            'line_chart_data': line_data,
            'repeated_users_data': rep_data,
            'busy_users_data': busy_data,
            'new_user_data': len(dept_new_users),
        }
    # print("request.session['depart'] - ", request.session['depart'])
    if request.session['depart'] == 'SuperAdmin':
        # print("inside - ", )
        return render(request, 'home/admin_dept_report.html', context)
    else:
        return render(request, 'home/department.html', context)


@csrf_protect
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_login(request):
    global admin_type, depart_name
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        dtype = request.POST.get('dtype')
        # print("dtype-> ", dtype)
        # print("==", type(dtype))

        user = authenticate(username=username, password=password)
        if user is not None:
            chk_super = User.objects.filter(username=username, is_superuser=True).exists()
            did = Department.objects.get(department=dtype)
            check_dpu = DepartmentAdminUser.objects.filter(user=user, department=did).exists()
            # print("check_dpu - ", check_dpu)
            # print("inside check super", did, dtype)
            if chk_super:
                # print("inside check super")
                login(request, user)
                request.session['depart'] = dtype
                request.session['admin_type'] = 'SuperAdmin'
                return redirect('department_data', department_name=dtype)
            elif check_dpu:
                user_ = DepartmentAdminUser.objects.get(user=user, department=did)
                admin_type = user_.usertype.usertype

                login(request, user)
                request.session['depart'] = dtype
                request.session['admin_type'] = admin_type
                return redirect('department_data', department_name=dtype)
            else:
                msg = f'YOU ARE NOT REGISTERED TO {dtype} DEPARTMENT'
                list_dept = Department.objects.all()
                context = {
                    'msg': msg,
                    'list_dept': list_dept
                }
                return render(request, 'accounts/login_.html', context)
        else:
            msg = 'INVALID CREDENTIALS'
            list_dept = Department.objects.all()
            context = {
                'msg': msg,
                'list_dept': list_dept
            }
            return render(request, 'accounts/login_.html', context)

    list_dept = Department.objects.all()
    context = {
        'list_dept': list_dept
    }
    return render(request, 'accounts/login_.html', context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def dept_logout(request):
    logout(request)
    return redirect('login_')


@login_required
def ExpoDptPDF(request, dept_name):
    wr_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='3').count()
    rt_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='4').count()
    no_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='5').count()
    live_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='6').count()
    reset_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='7').count()
    total_users_cnt = Log.objects.all().values('user_email').distinct().filter(intent=dept_name).count()

    _, _, rep_data, busy_data, new_user_data = dept_chart(dept_name)
    # print(">>> ", rep_data)
    # print(">type ", type(rep_data))
    # print(">type ", len(rep_data))
    print("###", busy_data)
    # print(">type ", type(busy_data))
    # print(">type ", len(busy_data))

    eng_users_cnt = total_users_cnt - len(new_user_data)
    context = {
        # 'dept_usr_list': dept_usr_list,
        'dept_name': dept_name,
        'wr_answer_count': wr_answer_cnt,
        'rt_answer_count': rt_answer_cnt,
        'no_answer_count': no_answer_cnt,
        'live_count': live_cnt,
        'reset_count': reset_cnt,
        'total_users': total_users_cnt,
        'eng_users_cnt': eng_users_cnt,
        'new_user_data': len(new_user_data),
        'repeated_users_list': rep_data,
        'busy_user_list': busy_data
    }

    pdf = render_to_pdf('home/dept_pdf_export.html', context)
    if pdf:
        return HttpResponse(pdf, content_type='application/pdf')
    return HttpResponse("PDF Not Found.")


@login_required
def ExpoDptExl(request, dept_name):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="DepartmentReport.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('DepartmentReport', cell_overwrite_ok=True)
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    columns = ['Total Users', 'Engaged Users', 'New Users', 'Reset Count', 'Live Chat', 'No Answer',
               'Right Answer', 'Wrong Answer']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    wr_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='3').count()
    rt_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='4').count()
    no_answer_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='5').count()
    live_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='6').count()
    reset_cnt = Log.objects.filter(intent=dept_name).filter(event_type_id='7').count()
    total_users_cnt = Log.objects.all().values('user_email').distinct().filter(intent=dept_name).count()

    _, _, rep_data, busy_data, new_user_data = dept_chart(dept_name)
    eng_users_cnt = total_users_cnt - len(new_user_data)

    rows_stats = [total_users_cnt, eng_users_cnt, len(new_user_data), reset_cnt,
                  live_cnt, no_answer_cnt, rt_answer_cnt, wr_answer_cnt]

    col_num_ = 0
    row_num += 1
    for row in range(len(rows_stats)):
        ws.write(row_num, col_num_, rows_stats[row], font_style)
        col_num_ += 1

    font_style.font.bold = True
    cols_tab = ['Table for Repeated Interacted Users Data']
    cols_repeated_users = ['User Email', 'Bot Interaction', 'Date Time']

    for col_num in range(len(cols_tab)):
        ws.write(3, col_num, cols_tab[col_num], font_style)

    for col_num in range(len(cols_repeated_users)):
        ws.write(4, col_num, cols_repeated_users[col_num], font_style)

    # rows_repeated_users = rep_data
    # print(">^^^^ ", rep_data)
    # print(">^^^^len ", len(rep_data))
    row_num += 3
    for row in rep_data:
        row_num += 1
        for row_ in row:
            # print("================= # =========================",row_)
            for col_num in range(len(cols_repeated_users)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)

    font_style.font.bold = True
    # busiest period table
    cols_tab = ['Table for Busiest period of Chatbot interaction']
    cols_busy_period_count = ['Department', 'Date Time', 'Count']

    for col_num in range(len(cols_tab)):
        ws.write(16, col_num, str(cols_tab[col_num]), font_style)

    for col_num in range(len(cols_busy_period_count)):
        ws.write(17, col_num, str(cols_busy_period_count[col_num]), font_style)

    # print("busy_data in exl", busy_data)
    row_num += 3
    for row in busy_data:
        row_num += 1
        for row_ in row:
            # print("================= # =========================",row_)
            for col_num in range(len(cols_repeated_users)):
                ws.write(row_num, col_num, str(row[col_num]), font_style)

    wb.save(response)

    if wb:
        return response
    return HttpResponse("No Data Found.")


def admin_check(user):
    if user.is_staff or user.is_superuser:
        return True
    return False


@login_required
def deptwise_usrlist(request):
    if request.session['admin_type'] == 'Department_User':
        return redirect('page_not_found')
    did = Department.objects.get(department=request.session['depart'])
    dept_usr_list = DepartmentAdminUser.objects.filter(department=did)

    # print("dept_usr_list >>> ", dept_usr_list)
    context = {
        'dept_usr_list': dept_usr_list,
        'depart_name': request.session['depart']
    }
    return render(request, 'home/dpt_usrlist.html', context)


def pdf_link_extract(pdf_name):
    link_regex = re.compile("((https?):((//)|(\\\\))+([\w\d:#@%/;$()~_?\+-=\\\.&](#!)?)*)",
                            re.DOTALL)  ## for pdf link extract
    text = extract_text(pdf_name)
    lst_links = re.findall(link_regex, text)
    return lst_links


def parse_jsonfile(json_path):
    res, keys = '', ''
    # print("parse_jsonfile_name-", json_path)
    # try:
    with open(json_path, 'r', encoding='utf-8') as f:
        # with open(filename, encoding='utf-8') as f:
        #     data = json.load(r'f')

        try:
            data = json.loads(f.read())
            # print("data - ", data)
            json_data = json.dumps(data)
            res = validateJSON(json_data)
            print(">>>", res)
            jsonData = data["assets"]
            print("in file ")
            for x in jsonData:
                keys = x.keys()
            return res, keys
        except:
            return False, None


def fetch_xml_tag(xml_file):
    xmlTree = ET.parse(xml_file)
    elemList = []
    for elem in xmlTree.iter():
        elemList.append(elem.tag)
    elemList = list(set(elemList))
    return elemList


def url_xml_val(url, xml_dir):
    url_nm = url.rsplit('/', 1)[1]
    data = requests.get(url)
    xml_file = xml_dir + os.sep + url_nm
    with open(xml_file, 'wb') as xml_:
        xml_.write(data.content)
        print("xml_file- ", xml_file)
        tag_lst = fetch_xml_tag(xml_file)
        try:
            parsefile(url)
            print("XML is well-formed", url)
        except:
            print("XML is NOT well-formed!", url)

    return url, tag_lst


def url_json_val(url):
    url_nm = url.rsplit('/', 1)[1]
    with urllib.request.urlopen(url) as url_link:
        data = json.loads(url_link.read().decode())
        json_data = json.dumps(data)
        result = validateJSON(json_data)
        jsonData = data["assets"]
        for x in jsonData:
            keys = x.keys()
        json_file = 'json_files' + os.sep + url_nm
        with open(json_file, 'w', encoding='utf-8') as json_:
            json.dump(data, json_)
    return result, jsonData, keys


# def UploadFile(request):
#     if request.method == 'POST':
#         form = UploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             return HttpResponseRedirect("/")
#     else:
#         form = UploadForm()
#         context = {
#             'form': form,
#         }
#     return render(request, 'home/Upload.html', context)

import shutil


def validate_url(request):
    global url_nm, result, ext_, url_path, get_tag, url
    if request.method == 'POST':
        url = request.POST.get('url')
        get_tag = request.POST.getlist('tag_lst')
        print(">>**", url)
        print("****", get_tag)
        result = ''

        ## fro file upload
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()

            # print("request.FILES - ", request.FILES['file'])
            # print("type  ", type(request.FILES['file']))

            ln = str(Upload.objects.last())
            print(">>", ln)
            ext_ = ln.split('.')[1]
            if ext_ == 'xml':
                print("xml file ", ln)

                try:
                    parsefile('media/' + ln)
                    tag_lst = fetch_xml_tag('media/' + ln)
                    print("XML is well-formed", ln)

                    context = {
                        'color': "green",
                        'msg': ln + '- is valid!',
                        'tag_lst': tag_lst,
                    }
                    return render(request, 'home/upload.html', context)

                except:
                    print("XML is NOT well-formed!", ln)

                    context = {
                        'color': "red",
                        'msg': ln + '- is invalid!',

                    }
                    return render(request, 'home/upload.html', context)

            elif ext_ == 'json':
                result, keys = parse_jsonfile('media/' + ln)
                if result:
                    context = {
                        'tag_lst': keys,
                        'color': 'green',
                        'msg': ln + '- is valid!',
                    }
                    return render(request, 'home/upload.html', context)
                else:
                    context = {
                        'color': 'red',
                        'msg': ln + '- is invalid!',
                    }
                return render(request, 'home/upload.html', context)

        if url:
            ext_ = url.split('.')[-1]

            if ext_ and url and ext_ == 'xml':

                # os.remove('xml_files')
                # os.mkdir('xml_files')
                url, tag_lst = url_xml_val(url, 'xml_files')
                context = {
                    'color': "green",
                    'msg': url + '- is valid!',
                    'tag_lst': tag_lst,
                    # 'msg1': 'from' + " " + url_nm + 'you have selected_tag(s)-' + " ".join(get_tag),
                }
                return render(request, 'home/upload.html', context)

            elif ext_ and url and ext_ == 'json':

                result, jsonData, keys = url_json_val(url)

                context = {
                    'tag_lst': keys,
                    'color': 'green',
                    'msg': url + '- is valid!',
                }
                return render(request, 'home/upload.html', context)

        elif get_tag:
            context = {
                'color': "green",
                'msg': 'you have selected a tag(s)-' + " ".join(get_tag),
            }
            return render(request, 'home/upload.html', context)

        else:
            form = UploadForm()
            context = {
                'color': "red",
                'msg': str(url) + '- is invalid!',
                'form': form,
            }
            return render(request, 'home/upload.html', context)

    return render(request, 'home/upload.html')
